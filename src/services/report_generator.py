# src/services/report_generator.py
# -*- coding: utf-8 -*-
"""
PDF・Excel レポート生成サービス
"""
from __future__ import annotations

import io
import os
from datetime import datetime
from typing import Any

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# Excel (openpyxl)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side,
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# PDF (reportlab)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 日本語フォント登録
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
_JP_FONT_NAME       = "HeiseiMin-W3"
_JP_FONT_REGISTERED = False


def _register_jp_font() -> str:
    """日本語フォントを登録し、使用するフォント名を返す"""
    global _JP_FONT_NAME, _JP_FONT_REGISTERED

    if _JP_FONT_REGISTERED:
        return _JP_FONT_NAME

    if not REPORTLAB_OK:
        return "Helvetica"

    # Windowsシステムフォント候補
    candidates = [
        r"C:\Windows\Fonts\msgothic.ttc",
        r"C:\Windows\Fonts\meiryo.ttc",
        r"C:\Windows\Fonts\YuGothM.ttc",
        # Linux (Streamlit Cloud)
        "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
        "/usr/share/fonts/opentype/noto/NotoSansCJKjp-Regular.otf",
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont("JPFont", path))
                _JP_FONT_NAME       = "JPFont"
                _JP_FONT_REGISTERED = True
                return _JP_FONT_NAME
            except Exception:
                continue

    # フォールバック: ReportLab内蔵CIDフォント
    try:
        from reportlab.pdfbase.cidfonts import UnicodeCIDFont
        pdfmetrics.registerFont(UnicodeCIDFont("HeiseiMin-W3"))
        _JP_FONT_NAME       = "HeiseiMin-W3"
        _JP_FONT_REGISTERED = True
    except Exception:
        _JP_FONT_NAME       = "Helvetica"
        _JP_FONT_REGISTERED = True

    return _JP_FONT_NAME


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# ReportGenerator クラス
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
class ReportGenerator:
    """
    PDF / Excel レポートを生成するサービスクラス。

    Parameters
    ----------
    case    : dict  ValuationCase の行データ（get_case の戻り値）
    params  : dict  ValuationParams フィールドのマッピング
    results : dict  ValuationResult フィールドのマッピング（複数モデル）
    greeks  : dict  Greeks 情報（省略可）
    """

    # カラーパレット
    COLOR_PRIMARY   = "#1E3A5F"
    COLOR_SECONDARY = "#2E86AB"
    COLOR_ACCENT    = "#F18F01"
    COLOR_BG_LIGHT  = "#F0F4F8"
    COLOR_SUCCESS   = "#27AE60"

    def __init__(
        self,
        case:    dict,
        params:  dict,
        results: dict,
        greeks:  dict | None = None,
    ) -> None:
        self.case    = case
        self.params  = params
        self.results = results
        self.greeks  = greeks or {}
        self._font   = _register_jp_font()

    # ────────────────────────────────────────
    # 公開 API
    # ────────────────────────────────────────
    def generate_pdf(self) -> bytes:
        """PDF バイト列を返す"""
        if not REPORTLAB_OK:
            raise ImportError("reportlab がインストールされていません")
        buf = io.BytesIO()
        self._build_pdf(buf)
        return buf.getvalue()

    def generate_excel(self) -> bytes:
        """Excel バイト列を返す"""
        if not OPENPYXL_OK:
            raise ImportError("openpyxl がインストールされていません")
        buf = io.BytesIO()
        self._build_excel(buf)
        return buf.getvalue()

    # ────────────────────────────────────────
    # PDF 内部処理
    # ────────────────────────────────────────
    def _build_pdf(self, buf: io.BytesIO) -> None:
        font = self._font
        doc  = SimpleDocTemplate(
            buf,
            pagesize=A4,
            leftMargin=20 * mm, rightMargin=20 * mm,
            topMargin=20 * mm,  bottomMargin=20 * mm,
        )

        st_title = ParagraphStyle(
            "Title_JP", fontName=font, fontSize=18, leading=24,
            textColor=colors.HexColor(self.COLOR_PRIMARY), spaceAfter=6,
        )
        st_h2 = ParagraphStyle(
            "H2_JP", fontName=font, fontSize=13, leading=18,
            textColor=colors.HexColor(self.COLOR_SECONDARY),
            spaceBefore=12, spaceAfter=4,
        )
        st_body = ParagraphStyle(
            "Body_JP", fontName=font, fontSize=9, leading=14,
        )
        st_small = ParagraphStyle(
            "Small_JP", fontName=font, fontSize=8, leading=12,
            textColor=colors.grey,
        )

        story: list[Any] = []

        # ── タイトル
        story.append(Paragraph("オプション評価レポート", st_title))
        story.append(Paragraph(
            f"ケース名: {self.case.get('case_name', '')} ／ "
            f"作成日時: {datetime.now().strftime('%Y年%m月%d日 %H:%M')}",
            st_body,
        ))
        story.append(HRFlowable(
            width="100%", thickness=2,
            color=colors.HexColor(self.COLOR_PRIMARY),
            spaceAfter=8,
        ))

        # ── 評価結果サマリー
        story.append(Paragraph("評価結果サマリー", st_h2))
        story.append(self._pdf_summary_table())
        story.append(Spacer(1, 8 * mm))

        # ── パラメータ
        story.append(Paragraph("入力パラメータ", st_h2))
        story.append(self._pdf_params_table())
        story.append(Spacer(1, 8 * mm))

        # ── モデル比較
        story.append(Paragraph("モデル別評価結果", st_h2))
        story.append(self._pdf_model_table())
        story.append(Spacer(1, 8 * mm))

        # ── Greeks
        if self.greeks:
            story.append(Paragraph("Greeks（感応度分析）", st_h2))
            story.append(self._pdf_greeks_table())
            story.append(Spacer(1, 8 * mm))

        # ── フッター注記
        story.append(HRFlowable(width="100%", thickness=1, color=colors.lightgrey))
        story.append(Spacer(1, 4 * mm))
        story.append(Paragraph(
            "※ 本レポートは参考資料であり、投資勧誘の趣旨とするものではありません。"
            "実際の評価は専門家に相談ください。",
            st_small,
        ))

        doc.build(story)

    def _pdf_table_style(self, header_color: str = "#1E3A5F") -> TableStyle:
        return TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  colors.HexColor(header_color)),
            ("TEXTCOLOR",      (0, 0), (-1, 0),  colors.white),
            ("FONTNAME",       (0, 0), (-1, -1), self._font),
            ("FONTSIZE",       (0, 0), (-1, 0),  10),
            ("FONTSIZE",       (0, 1), (-1, -1), 9),
            ("ALIGN",          (0, 0), (-1, -1), "CENTER"),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor(self.COLOR_BG_LIGHT)]),
            ("GRID",           (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING",     (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
        ])

    def _pdf_summary_table(self) -> Table:
        r = self.results
        wp = r.get("weighted_price")
        bs = r.get("bs_price")
        data = [
            ["項目",           "値"],
            ["加重平均価格",   f"¥{wp:,.4f}" if wp is not None else "未評価"],
            ["BS価格",         f"¥{bs:,.4f}" if bs is not None else "未評価"],
            ["オプション種類", r.get("option_type", "―")],
        ]
        tbl = Table(data, colWidths=[80 * mm, 80 * mm])
        tbl.setStyle(self._pdf_table_style(self.COLOR_PRIMARY))
        return tbl

    def _pdf_params_table(self) -> Table:
        p = self.params
        data = [
            ["パラメータ",              "値"],
            ["株価 (S)",                f"¥{p.get('stock_price', 0):,.1f}"],
            ["行使価格 (K)",            f"¥{p.get('strike_price', 0):,.1f}"],
            ["残存期間 (T)",            f"{p.get('time_to_expiry', 0):.4f} 年"],
            ["無リスク金利 (r)",        f"{p.get('risk_free_rate', 0) * 100:.2f}%"],
            ["ボラティリティ (σ)",      f"{p.get('volatility', 0) * 100:.2f}%"],
            ["配当利回り (q)",          f"{p.get('dividend_yield', 0) * 100:.2f}%"],
            ["二項ステップ数",          str(p.get('binomial_steps', 100))],
            ["MCシミュレーション数",    f"{p.get('mc_simulations', 10000):,}"],
        ]
        tbl = Table(data, colWidths=[80 * mm, 80 * mm])
        tbl.setStyle(self._pdf_table_style(self.COLOR_SECONDARY))
        return tbl

    def _pdf_model_table(self) -> Table:
        r = self.results
        data = [["モデル", "評価額", "ウェイト"]]
        model_map = [
            ("ブラック・ショールズ", "bs_price",       "50%"),
            ("二項モデル",           "binomial_price", "30%"),
            ("モンテカルロ",         "mc_price",       "20%"),
        ]
        for label, key, weight in model_map:
            val = r.get(key)
            data.append([label, f"¥{val:,.4f}" if val is not None else "―", weight])

        wp = r.get("weighted_price")
        data.append([
            "加重平均（最終評価額）",
            f"¥{wp:,.4f}" if wp is not None else "―",
            "―",
        ])

        tbl = Table(data, colWidths=[60 * mm, 70 * mm, 30 * mm])
        style = self._pdf_table_style(self.COLOR_SECONDARY)
        style.add("BACKGROUND", (0, -1), (-1, -1), colors.HexColor(self.COLOR_ACCENT))
        style.add("TEXTCOLOR",  (0, -1), (-1, -1), colors.white)
        style.add("FONTSIZE",   (0, -1), (-1, -1), 10)
        tbl.setStyle(style)
        return tbl

    def _pdf_greeks_table(self) -> Table:
        g = self.greeks
        data = [
            ["Greeks",    "値",                        "説明"],
            ["Delta (Δ)", f"{g.get('delta', 0):.4f}",  "株価1単位変化時のオプション価値変化"],
            ["Gamma (Γ)", f"{g.get('gamma', 0):.6f}",  "Deltaの変化率"],
            ["Theta (Θ)", f"{g.get('theta', 0):.4f}",  "1日経過によるオプション価値変化"],
            ["Vega (ν)",  f"{g.get('vega', 0):.4f}",   "ボラティリティ1%変化時の変化"],
            ["Rho (ρ)",   f"{g.get('rho', 0):.4f}",    "金利1%変化時の変化"],
        ]
        tbl = Table(data, colWidths=[35 * mm, 35 * mm, 90 * mm])
        tbl.setStyle(self._pdf_table_style(self.COLOR_PRIMARY))
        return tbl

    # ────────────────────────────────────────
    # Excel 内部処理
    # ────────────────────────────────────────
    def _build_excel(self, buf: io.BytesIO) -> None:
        wb = Workbook()
        wb.remove(wb.active)  # デフォルトシート削除

        self._excel_summary_sheet(wb)
        self._excel_params_sheet(wb)
        self._excel_model_sheet(wb)
        if self.greeks:
            self._excel_greeks_sheet(wb)

        wb.save(buf)

    # ── スタイル定数 ─────────────────────────
    @staticmethod
    def _hdr_fill(hex_color: str) -> PatternFill:
        return PatternFill("solid", fgColor=hex_color.lstrip("#"))

    @staticmethod
    def _thin_border() -> Border:
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    @staticmethod
    def _set_col_width(ws: Any, col: int, width: float) -> None:
        ws.column_dimensions[get_column_letter(col)].width = width

    def _write_header_row(
        self, ws: Any, row: int, headers: list[str], bg: str = "1E3A5F",
    ) -> None:
        for ci, hdr in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=ci, value=hdr)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = self._thin_border()

    def _write_data_row(
        self, ws: Any, row: int, values: list[Any], bg: str = "FFFFFF",
    ) -> None:
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True)
            cell.border    = self._thin_border()

    def _write_title(self, ws: Any, title: str, ncols: int) -> None:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1,   end_column=ncols,
        )
        cell = ws.cell(row=1, column=1, value=title)
        cell.font      = Font(bold=True, size=14, color="1E3A5F")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

    # ── シート別処理 ─────────────────────────
    def _excel_summary_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("評価サマリー")
        self._write_title(ws, "オプション評価レポート サマリー", 3)

        ws.cell(
            row=2, column=1,
            value=(
                f"ケース名: {self.case.get('case_name', '')}  ／  "
                f"作成日: {datetime.now().strftime('%Y/%m/%d %H:%M')}"
            ),
        ).font = Font(size=9, color="666666")
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=3)

        self._write_header_row(ws, 4, ["項目", "値", "備考"])

        r   = self.results
        wp  = r.get("weighted_price")
        bs  = r.get("bs_price")
        rows_data = [
            ("加重平均価格",   f"¥{wp:,.4f}" if wp is not None else "未評価", "最終評価額"),
            ("BS価格",         f"¥{bs:,.4f}" if bs is not None else "未評価", ""),
            ("オプション種類", r.get("option_type", "―"),                       ""),
        ]
        bgs = ["FFFFFF", "F0F4F8"]
        for i, row_vals in enumerate(rows_data):
            self._write_data_row(ws, 5 + i, list(row_vals), bgs[i % 2])

        for ci, w in enumerate([30, 25, 30], start=1):
            self._set_col_width(ws, ci, w)

    def _excel_params_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("入力パラメータ")
        self._write_title(ws, "入力パラメータ", 2)
        self._write_header_row(ws, 3, ["パラメータ", "値"], "2E86AB")

        p = self.params
        rows_data = [
            ("株価 (S)",             f"¥{p.get('stock_price', 0):,.1f}"),
            ("行使価格 (K)",         f"¥{p.get('strike_price', 0):,.1f}"),
            ("残存期間 (T)",         f"{p.get('time_to_expiry', 0):.4f} 年"),
            ("無リスク金利 (r)",     f"{p.get('risk_free_rate', 0) * 100:.2f}%"),
            ("ボラティリティ (σ)",   f"{p.get('volatility', 0) * 100:.2f}%"),
            ("配当利回り (q)",       f"{p.get('dividend_yield', 0) * 100:.2f}%"),
            ("二項ステップ数",       str(p.get('binomial_steps', 100))),
            ("MCシミュレーション数", f"{p.get('mc_simulations', 10000):,}"),
        ]
        bgs = ["FFFFFF", "F0F4F8"]
        for i, (k, v) in enumerate(rows_data):
            self._write_data_row(ws, 4 + i, [k, v], bgs[i % 2])

        self._set_col_width(ws, 1, 35)
        self._set_col_width(ws, 2, 25)

    def _excel_model_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("モデル比較")
        self._write_title(ws, "モデル別評価結果", 3)
        self._write_header_row(ws, 3, ["モデル", "評価額", "ウェイト"], "2E86AB")

        r = self.results
        model_map = [
            ("ブラック・ショールズ", "bs_price",       "50%"),
            ("二項モデル",           "binomial_price", "30%"),
            ("モンテカルロ",         "mc_price",       "20%"),
        ]
        bgs = ["FFFFFF", "F0F4F8"]
        for i, (label, key, weight) in enumerate(model_map):
            val = r.get(key)
            val_str = f"¥{val:,.4f}" if val is not None else "―"
            self._write_data_row(ws, 4 + i, [label, val_str, weight], bgs[i % 2])

        # 加重平均（最終評価額）行
        wp  = r.get("weighted_price")
        row = 4 + len(model_map)
        for ci, val in enumerate(
            ["加重平均（最終評価額）", f"¥{wp:,.4f}" if wp is not None else "―", "―"],
            start=1,
        ):
            cell = ws.cell(row=row, column=ci, value=val)
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill      = PatternFill("solid", fgColor="F18F01")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border    = self._thin_border()

        for ci, w in enumerate([30, 25, 15], start=1):
            self._set_col_width(ws, ci, w)

    def _excel_greeks_sheet(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Greeks")
        self._write_title(ws, "Greeks（感応度分析）", 3)
        self._write_header_row(ws, 3, ["Greeks", "値", "説明"], "1E3A5F")

        g = self.greeks
        rows_data = [
            ("Delta (Δ)", f"{g.get('delta', 0):.6f}", "株価1単位変化時のオプション価値変化"),
            ("Gamma (Γ)", f"{g.get('gamma', 0):.8f}", "Deltaの変化率"),
            ("Theta (Θ)", f"{g.get('theta', 0):.6f}", "1日経過によるオプション価値変化"),
            ("Vega (ν)",  f"{g.get('vega', 0):.6f}",  "ボラティリティ1%変化時の変化"),
            ("Rho (ρ)",   f"{g.get('rho', 0):.6f}",   "金利1%変化時の変化"),
        ]
        bgs = ["FFFFFF", "F0F4F8"]
        for i, row_vals in enumerate(rows_data):
            self._write_data_row(ws, 4 + i, list(row_vals), bgs[i % 2])

        for ci, w in enumerate([18, 20, 45], start=1):
            self._set_col_width(ws, ci, w)
